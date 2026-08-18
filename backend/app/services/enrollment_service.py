"""Reservations : prise de place, liste d'attente, recopie Google, export.

Ce module porte le systeme de reservation de la salle. Depuis qu'il collecte
aussi la formule de paiement et le montant encaisse, il tient le role du
registre : d'ou l'export Excel exhaustif que le back-office propose.
"""

from datetime import date as Date
from io import BytesIO
from uuid import uuid4

from fastapi import HTTPException, status
from sqlalchemy.orm import Session, joinedload

from app.models.models import Enrollment, ScheduleSlot
from app.services import google_form_service
from app.services.google_form_service import EnvoiGoogleEchoue


def _get_max_capacity(slot: ScheduleSlot) -> int:
    if slot.max_capacity_override is not None:
        return slot.max_capacity_override
    return slot.activity.max_capacity


def get_slot_availability(db: Session, slot_id, date=None) -> dict:
    slot = db.query(ScheduleSlot).filter(ScheduleSlot.id == slot_id).first()
    if not slot:
        return {"enrolled_count": 0, "max_capacity": 0, "available": 0}
    enrolled_count = (
        db.query(Enrollment)
        .filter(
            Enrollment.slot_id == slot_id,
            Enrollment.specific_date == date,
            Enrollment.status == "enrolled",
        )
        .count()
    )
    max_cap = _get_max_capacity(slot)
    return {
        "enrolled_count": enrolled_count,
        "max_capacity": max_cap,
        "available": max(0, max_cap - enrolled_count),
    }


def enroll(db: Session, data) -> Enrollment:
    # Sans ce controle, un slot_id inconnu part en base et la violation de cle
    # etrangere remonte en 500 au lieu du 404 attendu.
    slot = db.query(ScheduleSlot).filter(ScheduleSlot.id == data.slot_id).first()
    if slot is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Creneau introuvable",
        )

    availability = get_slot_availability(db, data.slot_id, data.specific_date)
    # Nom distinct de `status` : ce dernier est le module importe de fastapi,
    # une variable locale du meme nom le masquerait dans toute la fonction.
    statut = "enrolled" if availability["available"] > 0 else "waitlisted"
    enrollment = Enrollment(
        id=str(uuid4()),
        user_name=data.user_name,
        user_email=data.user_email,
        user_phone=data.user_phone,
        slot_id=data.slot_id,
        specific_date=data.specific_date,
        status=statut,
        session_type=data.session_type,
        payment_type=data.payment_type,
        amount_paid=data.amount_paid,
        feedback=data.feedback,
    )
    db.add(enrollment)
    db.commit()
    db.refresh(enrollment)

    recopier_vers_google(db, enrollment)
    return enrollment


def cancel_enrollment(db: Session, enrollment_id) -> bool:
    enrollment = db.query(Enrollment).filter(Enrollment.id == enrollment_id).first()
    if not enrollment:
        return False
    enrollment.status = "cancelled"
    db.commit()

    _promote_waitlisted(db, enrollment.slot_id, enrollment.specific_date)
    return True


def _promote_waitlisted(db: Session, slot_id, specific_date):
    availability = get_slot_availability(db, slot_id, specific_date)
    if availability["available"] <= 0:
        return
    next_waitlisted = (
        db.query(Enrollment)
        .filter(
            Enrollment.slot_id == slot_id,
            Enrollment.specific_date == specific_date,
            Enrollment.status == "waitlisted",
        )
        .order_by(Enrollment.enrolled_at)
        .first()
    )
    if next_waitlisted:
        next_waitlisted.status = "enrolled"
        db.commit()


def get_slot_enrollments(db: Session, slot_id, date=None) -> list[Enrollment]:
    query = db.query(Enrollment).filter(
        Enrollment.slot_id == slot_id,
        Enrollment.status != "cancelled",
    )
    if date is not None:
        query = query.filter(Enrollment.specific_date == date)
    return query.order_by(Enrollment.enrolled_at).all()


def recopier_vers_google(db: Session, enrollment: Enrollment) -> Enrollment:
    """Recopie la reservation dans le formulaire Google, sans jamais echouer.

    L'ordre compte : la place est deja prise quand on arrive ici. Si Google
    refuse ou ne repond pas, le visiteur ne doit pas voir d'erreur pour autant.
    L'echec est note sur la ligne, le back-office le signale et permet de le
    rejouer.
    """
    try:
        google_form_service.envoyer(enrollment)
        enrollment.forwarded_to_google = True
        enrollment.google_error = None
    except EnvoiGoogleEchoue as erreur:
        enrollment.forwarded_to_google = False
        enrollment.google_error = str(erreur)
    db.commit()
    db.refresh(enrollment)
    return enrollment


def resend_to_google(db: Session, enrollment_id) -> Enrollment | None:
    enrollment = _requete_complete(db).filter(Enrollment.id == enrollment_id).first()
    if not enrollment:
        return None
    return recopier_vers_google(db, enrollment)


def _requete_complete(db: Session):
    """Reservations avec leur creneau, son activite et son coach.

    Sans ces `joinedload`, la liste et l'export declenchent trois requetes par
    ligne — et l'export d'une annee entiere devient interminable.
    """
    return db.query(Enrollment).options(
        joinedload(Enrollment.slot).joinedload(ScheduleSlot.activity),
        joinedload(Enrollment.slot).joinedload(ScheduleSlot.coach),
    )


def get_all_enrollments(
    db: Session,
    depuis: Date | None = None,
    jusqu_a: Date | None = None,
    statut: str | None = None,
    payment_type: str | None = None,
    session_type: str | None = None,
    activity_id: str | None = None,
) -> list[Enrollment]:
    """Toutes les reservations, filtrables. Reservee au back-office."""
    query = _requete_complete(db)
    if depuis:
        query = query.filter(Enrollment.specific_date >= depuis)
    if jusqu_a:
        query = query.filter(Enrollment.specific_date <= jusqu_a)
    if statut:
        query = query.filter(Enrollment.status == statut)
    if payment_type:
        query = query.filter(Enrollment.payment_type == payment_type)
    if session_type:
        query = query.filter(Enrollment.session_type == session_type)
    if activity_id:
        query = query.join(ScheduleSlot).filter(ScheduleSlot.activity_id == activity_id)
    return query.order_by(
        Enrollment.specific_date.desc(), Enrollment.enrolled_at.desc()
    ).all()


STATUTS_EN_CLAIR = {
    "enrolled": "Inscrit",
    "waitlisted": "Liste d'attente",
    "cancelled": "Annule",
}


def _valeur_exportee(enrollment: Enrollment, cle: str):
    """Une cellule du classeur. Aplatit le creneau, l'activite et le coach."""
    slot = enrollment.slot
    if cle == "activite":
        return slot.activity.name if slot and slot.activity else None
    if cle == "coach":
        return slot.coach.name if slot and slot.coach else None
    if cle == "horaire":
        return f"{slot.start_time} - {slot.end_time}" if slot else None
    if cle == "statut":
        return STATUTS_EN_CLAIR.get(enrollment.status, enrollment.status)
    if cle == "forwarded_to_google":
        return "Oui" if enrollment.forwarded_to_google else "Non"
    if cle == "amount_paid":
        return None if enrollment.amount_paid is None else float(enrollment.amount_paid)
    if cle == "enrolled_at":
        return enrollment.enrolled_at.replace(tzinfo=None) if enrollment.enrolled_at else None
    return getattr(enrollment, cle, None)


# Colonnes de l'export, dans l'ordre. Exhaustif au sens propre : tout ce que la
# base retient d'une reservation s'y retrouve, y compris ce qui est calcule
# depuis le creneau.
COLONNES = [
    ("Date de seance", "specific_date", 15),
    ("Horaire", "horaire", 16),
    ("Activite", "activite", 22),
    ("Coach", "coach", 18),
    ("Nom du membre", "user_name", 26),
    ("E-mail", "user_email", 30),
    ("Telephone", "user_phone", 18),
    ("Statut", "statut", 16),
    ("Type de seance", "session_type", 16),
    ("Formule de paiement", "payment_type", 22),
    ("Montant paye (FCFA)", "amount_paid", 20),
    ("Remarque du membre", "feedback", 44),
    ("Recopie Google", "forwarded_to_google", 15),
    ("Reserve le", "enrolled_at", 20),
]


def export_xlsx(enrollments: list[Enrollment]) -> bytes:
    """Classeur Excel des reservations, une ligne par reservation.

    `openpyxl` ecrit un vrai `.xlsx` : un CSV renomme s'ouvrirait de travers
    dans Excel des qu'une remarque contient une virgule ou un retour a la ligne.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Reservations"

    entete_fond = PatternFill("solid", fgColor="0F1724")
    entete_police = Font(color="FFD600", bold=True)

    for numero, (libelle, _, largeur) in enumerate(COLONNES, start=1):
        cellule = feuille.cell(row=1, column=numero, value=libelle)
        cellule.fill = entete_fond
        cellule.font = entete_police
        cellule.alignment = Alignment(vertical="center")
        feuille.column_dimensions[get_column_letter(numero)].width = largeur

    for ligne, enrollment in enumerate(enrollments, start=2):
        for numero, (_, cle, _) in enumerate(COLONNES, start=1):
            feuille.cell(row=ligne, column=numero, value=_valeur_exportee(enrollment, cle))

    # Fige l'en-tete et pose le filtre automatique : le fichier s'ouvre
    # directement utilisable, sans manipulation.
    feuille.freeze_panes = "A2"
    feuille.auto_filter.ref = f"A1:{get_column_letter(len(COLONNES))}{max(1, len(enrollments) + 1)}"

    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
