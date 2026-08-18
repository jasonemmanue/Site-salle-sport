"""Le registre des reservations : renseignements de paiement, recopie, export.

Notre systeme de reservation collecte desormais ce que le registre de la salle
notait a la main : type de seance, formule de paiement, montant encaisse,
remarque du membre. Ces informations sont facultatives — une reservation reste
valable sans elles — mais leur presence conditionne la recopie vers le
formulaire Google, qui les exige.

Deux regles structurent ce fichier :

- **la reservation fait foi.** Un echec de recopie ne l'annule jamais, il est
  seulement note sur la ligne ;
- **l'export dit tout.** Ce que la base retient d'une reservation doit se
  retrouver dans le classeur, y compris ce qui vient du creneau — activite,
  horaire, coach.

Aucun test ne poste dans le vrai formulaire : `GOOGLE_FORM_ENABLED` est a faux
dans `conftest.py`, et les tests qui portent sur la recopie remplacent
`google_form_service.envoyer`.
"""

from datetime import date, timedelta
from io import BytesIO

import pytest

from app.services import enrollment_service, google_form_service
from app.services.google_form_service import EnvoiGoogleEchoue
from tests.conftest import UUID_ABSENT

CHEMIN = "/api/v1/enrollments"
TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

LUNDI = date.today() - timedelta(days=date.today().weekday())


def _reserver(client, creneau, prenom="Ama", **surcharges):
    corps = {
        "user_name": f"{prenom} Kouassi",
        "user_email": f"{prenom.lower()}@tests.eslie",
        "user_phone": "+225 0545079850",
        "slot_id": creneau.id,
        "specific_date": LUNDI.isoformat(),
        "session_type": "Collectif",
        "payment_type": "Séance",
        "amount_paid": 3000,
        "feedback": "Premiere seance, tres bon accueil.",
    }
    corps.update(surcharges)
    return client.post(f"{CHEMIN}/", json=corps)


@pytest.fixture
def recopie_reussie(monkeypatch):
    envoyees = []
    monkeypatch.setattr(google_form_service, "envoyer", lambda e: envoyees.append(e))
    return envoyees


@pytest.fixture
def recopie_en_panne(monkeypatch):
    def echouer(_):
        raise EnvoiGoogleEchoue("Google injoignable : delai depasse")

    monkeypatch.setattr(google_form_service, "envoyer", echouer)


# --- Renseignements complementaires ---------------------------------------


def test_les_renseignements_de_paiement_sont_conserves(client, creneau, recopie_reussie):
    corps = _reserver(client, creneau).json()
    assert corps["session_type"] == "Collectif"
    assert corps["payment_type"] == "Séance"
    assert corps["amount_paid"] == 3000.0
    assert corps["feedback"] == "Premiere seance, tres bon accueil."


def test_une_reservation_sans_renseignements_reste_valable(client, creneau, recopie_reussie):
    """Ils sont facultatifs : le parcours d'avant continue de fonctionner."""
    reponse = client.post(
        f"{CHEMIN}/",
        json={
            "user_name": "Ama",
            "user_email": "ama@tests.eslie",
            "user_phone": "0000",
            "slot_id": creneau.id,
            "specific_date": LUNDI.isoformat(),
        },
    )
    assert reponse.status_code == 201
    assert reponse.json()["status"] == "enrolled"
    assert reponse.json()["payment_type"] is None


def test_type_de_seance_hors_liste(client, creneau):
    assert _reserver(client, creneau, session_type="Duo").status_code == 422


def test_formule_de_paiement_sans_accent_refusee(client, creneau):
    """« Seance » n'est pas « Séance » : Google refuserait, on refuse avant lui."""
    assert _reserver(client, creneau, payment_type="Seance").status_code == 422


def test_les_quatre_formules_sont_acceptees(client, creneau, recopie_reussie):
    for formule in (
        "Abonnée mensuel",
        "Séance",
        "Abonnement de karaté",
        "Abonnement de box",
    ):
        assert _reserver(client, creneau, payment_type=formule).status_code == 201, formule


def test_montant_negatif_refuse(client, creneau):
    assert _reserver(client, creneau, amount_paid=-500).status_code == 422


# --- Recopie vers le formulaire Google -------------------------------------


def test_la_recopie_est_marquee_quand_elle_aboutit(client, creneau, recopie_reussie):
    corps = _reserver(client, creneau).json()
    assert corps["forwarded_to_google"] is True
    assert corps["google_error"] is None
    assert len(recopie_reussie) == 1


def test_un_echec_de_recopie_n_annule_pas_la_reservation(client, db, creneau, recopie_en_panne):
    """Le point central : la place est prise meme si Google ne repond pas."""
    from app.models.models import Enrollment

    reponse = _reserver(client, creneau)
    assert reponse.status_code == 201
    corps = reponse.json()
    assert corps["status"] == "enrolled"
    assert corps["forwarded_to_google"] is False
    assert "injoignable" in corps["google_error"]
    assert db.query(Enrollment).filter(Enrollment.id == corps["id"]).first() is not None


def test_sans_renseignements_la_recopie_est_annoncee_impossible(client, creneau, monkeypatch):
    """Google exige la formule et le type : on le dit, au lieu d'essuyer un 400."""
    monkeypatch.setattr(google_form_service.settings, "GOOGLE_FORM_ENABLED", True)
    reponse = client.post(
        f"{CHEMIN}/",
        json={
            "user_name": "Ama",
            "user_email": "ama@tests.eslie",
            "user_phone": "0000",
            "slot_id": creneau.id,
            "specific_date": LUNDI.isoformat(),
        },
    )
    corps = reponse.json()
    assert corps["forwarded_to_google"] is False
    assert "Type de seance" in corps["google_error"]


def test_la_charge_envoyee_porte_les_identifiants_du_formulaire(client, db, creneau, recopie_reussie):
    _reserver(client, creneau)
    enrollment = recopie_reussie[0]
    charge = google_form_service.construire_charge(enrollment)
    assert charge["entry.1859557394"] == LUNDI.isoformat()
    assert charge["entry.1963678662"] == "Ama Kouassi"
    assert charge["entry.1062618537"] == "Collectif"
    assert charge["entry.1243048007"] == "Séance"
    assert charge["entry.1438213460"] == "3000"
    # Le nom du coach vient du creneau : le membre n'a pas a le ressaisir.
    assert charge["entry.1231872317"] == "Toussaint"


def test_la_date_part_en_une_seule_valeur(client, creneau, recopie_reussie):
    """Le triplet `_year`/`_month`/`_day` fait repondre 400 a ce formulaire."""
    _reserver(client, creneau)
    charge = google_form_service.construire_charge(recopie_reussie[0])
    assert not [cle for cle in charge if cle.endswith(("_year", "_month", "_day"))]


def test_les_champs_vides_ne_sont_pas_envoyes(client, creneau, recopie_reussie):
    _reserver(client, creneau, amount_paid=None, feedback=None)
    charge = google_form_service.construire_charge(recopie_reussie[0])
    assert "entry.1438213460" not in charge
    assert "entry.1797340128" not in charge


def test_rejouer_la_recopie_apres_un_echec(client, monkeypatch, entetes_admin, creneau, recopie_en_panne):
    identifiant = _reserver(client, creneau).json()["id"]

    monkeypatch.setattr(google_form_service, "envoyer", lambda e: None)
    reponse = client.post(f"{CHEMIN}/{identifiant}/resend", headers=entetes_admin)
    assert reponse.status_code == 200
    assert reponse.json()["forwarded_to_google"] is True
    assert reponse.json()["google_error"] is None


def test_rejouer_reserve_a_l_admin(client, creneau, recopie_reussie):
    identifiant = _reserver(client, creneau).json()["id"]
    assert client.post(f"{CHEMIN}/{identifiant}/resend").status_code == 403


def test_rejouer_identifiant_absent(client, entetes_admin):
    assert client.post(f"{CHEMIN}/{UUID_ABSENT}/resend", headers=entetes_admin).status_code == 404


# --- Liste du back-office --------------------------------------------------


def test_la_liste_complete_est_reservee_a_l_admin(client):
    assert client.get(f"{CHEMIN}/").status_code == 403


def test_la_liste_complete_est_refusee_a_un_membre(client, membre):
    assert client.get(f"{CHEMIN}/", headers=membre.entetes).status_code == 403


def test_la_liste_porte_le_creneau_son_activite_et_son_coach(client, entetes_admin, creneau, recopie_reussie):
    _reserver(client, creneau)
    ligne = client.get(f"{CHEMIN}/", headers=entetes_admin).json()[0]
    assert ligne["slot"]["activity"]["name"] == "Musculation"
    assert ligne["slot"]["coach"]["name"] == "Toussaint"
    assert ligne["slot"]["start_time"] == "07:00"


def test_la_liste_montre_aussi_les_annulations(client, entetes_admin, creneau, recopie_reussie):
    """Contrairement a la vue par creneau : un registre ne s'ampute pas."""
    identifiant = _reserver(client, creneau).json()["id"]
    client.delete(f"{CHEMIN}/{identifiant}")
    lignes = client.get(f"{CHEMIN}/", headers=entetes_admin).json()
    assert [ligne["status"] for ligne in lignes] == ["cancelled"]


def test_filtre_par_periode(client, entetes_admin, creneau, recopie_reussie):
    _reserver(client, creneau, "Ama", specific_date=(LUNDI - timedelta(days=10)).isoformat())
    _reserver(client, creneau, "Bakary", specific_date=LUNDI.isoformat())
    reponse = client.get(
        f"{CHEMIN}/", params={"depuis": LUNDI.isoformat()}, headers=entetes_admin
    )
    assert len(reponse.json()) == 1


def test_filtre_par_formule_de_paiement(client, entetes_admin, creneau, recopie_reussie):
    _reserver(client, creneau, "Ama", payment_type="Séance")
    _reserver(client, creneau, "Bakary", payment_type="Abonnement de box")
    reponse = client.get(
        f"{CHEMIN}/", params={"payment_type": "Abonnement de box"}, headers=entetes_admin
    )
    assert len(reponse.json()) == 1


def test_filtre_par_statut(client, entetes_admin, creneau, recopie_reussie):
    identifiant = _reserver(client, creneau, "Ama").json()["id"]
    _reserver(client, creneau, "Bakary")
    client.delete(f"{CHEMIN}/{identifiant}")
    reponse = client.get(f"{CHEMIN}/", params={"statut": "cancelled"}, headers=entetes_admin)
    assert len(reponse.json()) == 1


def test_filtre_par_activite(client, entetes_admin, creneau, activite, recopie_reussie):
    _reserver(client, creneau)
    assert len(client.get(f"{CHEMIN}/", params={"activity_id": activite.id}, headers=entetes_admin).json()) == 1
    assert client.get(f"{CHEMIN}/", params={"activity_id": UUID_ABSENT}, headers=entetes_admin).json() == []


def test_tri_de_la_seance_la_plus_recente_a_la_plus_ancienne(client, entetes_admin, creneau, recopie_reussie):
    _reserver(client, creneau, "Ancienne", specific_date=(LUNDI - timedelta(days=7)).isoformat())
    _reserver(client, creneau, "Recente", specific_date=LUNDI.isoformat())
    noms = [l["user_name"] for l in client.get(f"{CHEMIN}/", headers=entetes_admin).json()]
    assert noms == ["Recente Kouassi", "Ancienne Kouassi"]


# --- Export Excel ----------------------------------------------------------


def test_export_reserve_a_l_admin(client):
    assert client.get(f"{CHEMIN}/export.xlsx").status_code == 403


def test_export_refuse_a_un_membre(client, membre):
    assert client.get(f"{CHEMIN}/export.xlsx", headers=membre.entetes).status_code == 403


def test_l_export_est_un_vrai_classeur_excel(client, entetes_admin, creneau, recopie_reussie):
    _reserver(client, creneau)
    reponse = client.get(f"{CHEMIN}/export.xlsx", headers=entetes_admin)
    assert reponse.status_code == 200
    assert reponse.headers["content-type"] == TYPE_XLSX
    assert reponse.headers["content-disposition"].startswith("attachment;")
    # Un .xlsx est une archive ZIP : la signature le prouve mieux qu'un
    # `content-type`, qu'on pourrait poser sur n'importe quoi.
    assert reponse.content[:2] == b"PK"


def test_le_classeur_dit_tout_de_la_reservation(client, entetes_admin, creneau, recopie_reussie):
    from openpyxl import load_workbook

    _reserver(client, creneau)
    contenu = client.get(f"{CHEMIN}/export.xlsx", headers=entetes_admin).content
    feuille = load_workbook(BytesIO(contenu)).active

    entetes = [c.value for c in feuille[1]]
    assert entetes == [libelle for libelle, _, _ in enrollment_service.COLONNES]

    ligne = {entetes[i]: c.value for i, c in enumerate(feuille[2])}
    assert ligne["Date de seance"].date() == LUNDI
    assert ligne["Horaire"] == "07:00 - 08:00"
    assert ligne["Activite"] == "Musculation"
    assert ligne["Coach"] == "Toussaint"
    assert ligne["Nom du membre"] == "Ama Kouassi"
    assert ligne["E-mail"] == "ama@tests.eslie"
    assert ligne["Telephone"] == "+225 0545079850"
    assert ligne["Statut"] == "Inscrit"
    assert ligne["Type de seance"] == "Collectif"
    assert ligne["Formule de paiement"] == "Séance"
    assert ligne["Montant paye (FCFA)"] == 3000
    assert ligne["Remarque du membre"] == "Premiere seance, tres bon accueil."
    assert ligne["Recopie Google"] == "Oui"


def test_le_statut_est_ecrit_en_clair(client, entetes_admin, creneau, recopie_reussie):
    """« cancelled » ne dit rien a la personne qui ouvre le fichier."""
    from openpyxl import load_workbook

    identifiant = _reserver(client, creneau).json()["id"]
    client.delete(f"{CHEMIN}/{identifiant}")
    contenu = client.get(f"{CHEMIN}/export.xlsx", headers=entetes_admin).content
    feuille = load_workbook(BytesIO(contenu)).active
    colonne = [l for l, _, _ in enrollment_service.COLONNES].index("Statut") + 1
    assert feuille.cell(row=2, column=colonne).value == "Annule"


def test_le_classeur_respecte_les_filtres(client, entetes_admin, creneau, recopie_reussie):
    from openpyxl import load_workbook

    _reserver(client, creneau, "Ama", payment_type="Séance")
    _reserver(client, creneau, "Bakary", payment_type="Abonnement de box")
    contenu = client.get(
        f"{CHEMIN}/export.xlsx", params={"payment_type": "Séance"}, headers=entetes_admin
    ).content
    assert load_workbook(BytesIO(contenu)).active.max_row == 2


def test_un_export_vide_garde_son_entete(client, entetes_admin):
    from openpyxl import load_workbook

    contenu = client.get(f"{CHEMIN}/export.xlsx", headers=entetes_admin).content
    feuille = load_workbook(BytesIO(contenu)).active
    assert feuille.max_row == 1
    assert feuille["A1"].value == "Date de seance"


def test_l_entete_est_fige_et_filtrable(client, entetes_admin, creneau, recopie_reussie):
    """Confort d'usage : le fichier s'ouvre pret a etre trie."""
    from openpyxl import load_workbook

    _reserver(client, creneau)
    contenu = client.get(f"{CHEMIN}/export.xlsx", headers=entetes_admin).content
    feuille = load_workbook(BytesIO(contenu)).active
    assert feuille.freeze_panes == "A2"
    assert feuille.auto_filter.ref.startswith("A1:")
